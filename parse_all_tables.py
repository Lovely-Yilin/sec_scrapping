#!/usr/bin/env python3
"""Extract all HTML tables from a single SEC filing URL using parser_new."""

import argparse
import json
import re
from pathlib import Path
from datetime import datetime

from parser import FILING_URL, HEADERS, parse_all_tables


def _safe_slug(text):
    text = (text or "").strip().lower()
    keep = []
    for ch in text:
        if ch.isalnum():
            keep.append(ch)
        elif ch in (" ", "-", "_"):
            keep.append("_")
    slug = "".join(keep)
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_") or "table"


def _make_sheet_name(index, heading, used_names):
    base = f"t{index:03d}_{_safe_slug(heading)}"
    # Excel worksheet names must be <= 31 chars and unique per workbook.
    base = base[:31] or f"t{index:03d}"
    candidate = base
    suffix_num = 1
    while candidate in used_names:
        suffix = f"_{suffix_num}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        suffix_num += 1
    used_names.add(candidate)
    return candidate


def _derive_output_stem(url):
    # Prefer the filing filename stem from URL, e.g. nflx-20251231.htm -> nflx-20251231.
    stem_match = re.search(r"/([^/?#]+)\.htm(?:[?#].*)?$", url or "", flags=re.IGNORECASE)
    if stem_match:
        stem = stem_match.group(1).strip()
        if stem:
            return stem

    # Fallback to date-based naming when URL does not contain a .htm filename.
    matches = re.findall(r"\b(20\d{6})\b", url or "")
    date_part = matches[-1] if matches else datetime.today().strftime("%Y%m%d")
    return f"filing-{date_part}"


def _write_all_tables_excel(tables, out_dir, url):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write Excel output. Install with: pip install openpyxl") from exc

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    file_stem = _derive_output_stem(url)
    excel_path = out_path / f"{file_stem}.xlsx"

    wb = Workbook()
    # Remove the default empty sheet so only parsed tables are present.
    default_ws = wb.active
    wb.remove(default_ws)

    used_sheet_names = set()
    written_count = 0
    for t in tables:
        columns = list(t.get("columns") or [])
        data_rows = [list(r) for r in (t.get("data_rows") or [])]
        if not columns:
            rows = t.get("rows") or []
            if not rows:
                continue
            width = max(len(r) for r in rows)
            columns = [f"Column {i + 1}" for i in range(width)]
            data_rows = [list(r) + [""] * (width - len(r)) for r in rows]

        if not columns and not data_rows:
            continue

        width = len(columns)
        padded_data = [r + [""] * (width - len(r)) for r in data_rows]
        sheet_name = _make_sheet_name(t["index"], t.get("heading", ""), used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns)
        for row in padded_data:
            ws.append(row)
        written_count += 1

    # If no table is usable, keep one sheet explaining the result.
    if not wb.worksheets:
        ws = wb.create_sheet(title="tables")
        ws.append(["No tables with writable rows were found."])

    # save() overwrites an existing file with the same name.
    wb.save(excel_path)
    return str(excel_path), written_count


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--url",
        default=FILING_URL,
        help="SEC filing .htm URL (default: parser_new.FILING_URL)",
    )
    ap.add_argument(
        "--output-json",
        default="",
        help="Optional path to write full extracted table payload as JSON.",
    )
    ap.add_argument(
        "--strict-csv-dir",
        default="",
        help="Optional directory to write a single Excel file with one sheet per table.",
    )
    args = ap.parse_args()

    tables = parse_all_tables(args.url, headers=HEADERS)
    print(f"Extracted {len(tables)} tables from {args.url}\n")
    print("=" * 80)
    for t in tables:
        print(f"\n--- Table {t['index']} ---")
        if t["heading"]:
            print(f"Heading: {t['heading']}")
        text = t["text"]
        print(text[:2000])
        if len(text) > 2000:
            print(f"  ... (truncated, full length {len(text)} chars)")
        print()

    if args.output_json:
        with open(args.output_json, "w", encoding="utf-8") as f:
            json.dump(tables, f, indent=2, ensure_ascii=False)
        print(f"Wrote full table payload to {args.output_json}")

    if args.strict_csv_dir:
        excel_path, count = _write_all_tables_excel(tables, args.strict_csv_dir, args.url)
        print(f"Wrote {count} tables to Excel file: {excel_path}")


if __name__ == "__main__":
    main()
